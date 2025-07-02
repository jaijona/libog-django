
from django.db.models.functions import TruncDate
from django.db.models import Avg, DateField
from datetime import datetime
from usuarios.utils import login_requerido
from .models import Promedio
from django.shortcuts import render
from django.core.paginator import Paginator
from modelos_app.models import ModeloRegistrado
import openpyxl
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from django.http import JsonResponse
from promedio.utils import guardar_promedios 
from django.utils.timezone import make_aware
from datetime import datetime,timedelta
from ingresos_tokens.models import IngresoTokens
from django.db.models import Sum

from django.core.management import call_command
from django.views.decorators.csrf import csrf_exempt
from django.conf import settings






import os


CLAVE_SECRETA = os.getenv("CLAVE_SECRETA")

def api_guardar_promedios(request):
    key = request.GET.get("key")

    if key != CLAVE_SECRETA:
        return JsonResponse({"error": "No autorizado"}, status=403)

    try:
        guardar_promedios()
        return JsonResponse({"status": "ok"})
    except Exception as e:
        return JsonResponse({"error": str(e)}, status=500)
    



@login_requerido
def ver_promedios(request):
    id_studio = request.session.get('id_studio')
    print("studio_id:", id_studio)  # ← Esto ahora debe aparecer
    datos = Promedio.objects.filter(id_studio=id_studio,id_modelo__estado=1).select_related('id_modelo')

    # Obtener filtros
    fecha_filtro = request.GET.get('fecha')
    usuario_filtro = request.GET.get('usuario')
    jornada_filtro = request.GET.get('jornada')
    
    """
    # Aplicar filtro por fecha
    if fecha_filtro:
        try:
            # Convertir el string a date
            colombia_tz = pytz.timezone('America/Bogota')
            fecha_base = datetime.strptime(fecha_filtro, '%Y-%m-%d')

            # Crear rango completo del día con zona horaria
            inicio = make_aware(datetime.combine(fecha_base, time.min),colombia_tz)  # 00:00:00
            fin = make_aware(datetime.combine(fecha_base, time.max),colombia_tz)     # 23:59:59.999999

            # Aplicar el filtro
            datos = datos.filter(fecha__range=(inicio, fin))
        except ValueError:
            pass
    """
    # Aplicar filtro por usuario
    if usuario_filtro:
        datos = datos.filter(id_modelo__usuario=usuario_filtro)

    # Aplicar filtro por jornada
    if jornada_filtro:
        datos = datos.filter(id_modelo__jornada=jornada_filtro)

    # Paginación (15 por página)
    paginator = Paginator(datos.order_by('-fecha'), 15)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    # Lista de usuarios para mostrar en el dropdown
    usuarios = ModeloRegistrado.objects.filter(studio_id=id_studio,estado=1).values_list('usuario', flat=True).distinct()
    jornada = ModeloRegistrado.objects.filter(studio_id=id_studio,estado=1).values_list('jornada', flat=True).distinct()

    context = {
        'page_obj': page_obj,
        'usuarios': usuarios,
        'jornada': jornada,
        'fecha_filtro': fecha_filtro or '',
        'usuario_filtro': usuario_filtro or '',
        'jornada_filtro': jornada_filtro or '',
    }
    return render(request, 'ver_promedios.html', context)

def exportar_excel(request):
    id_studio = request.session.get('usuario_id')
    datos = Promedio.objects.filter(id_studio=id_studio,id_modelo__estado=1).select_related('id_modelo')

    # Filtros
    fecha_filtro = request.GET.get('fecha')
    usuario_filtro = request.GET.get('usuario')
    jornada_filtro = request.GET.get('jornada')

    if fecha_filtro:
        try:
            fecha_obj = datetime.strptime(fecha_filtro, '%Y-%m-%d').date()
            datos = datos.filter(fecha=fecha_obj)
        except ValueError:
            pass

    if usuario_filtro:
        datos = datos.filter(id_modelo__usuario=usuario_filtro)


    if jornada_filtro:
        datos = datos.filter(id_modelo__jornada=jornada_filtro)

    # Crear workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Promedios"

    # Encabezados
    headers = ['Usuario', 'Jornada', 'Promedio','Usuarios', 'Fecha']
    ws.append(headers)

    for item in datos:
        ws.append([
            item.id_modelo.usuario,
            item.id_modelo.jornada,
            item.promedio,
            item.users,
            item.fecha.strftime('%Y-%m-%d')
        ])

    # Ajustar ancho de columnas
    for i, column in enumerate(ws.columns, 1):
        max_length = max(len(str(cell.value)) for cell in column)
        ws.column_dimensions[get_column_letter(i)].width = max_length + 2

    # Respuesta HTTP
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=promedios_filtrados.xlsx'
    wb.save(response)
    return response


####
####
# INICIO DE FUNCIONES DE GRAFICACION
####
####

def obtener_datos_semana(request):
    modelo_id = request.GET.get('modelo')
    fecha_inicio_str = request.GET.get('fecha_inicio')

    if not (modelo_id and fecha_inicio_str):
        return JsonResponse({'error': 'Faltan parámetros'}, status=400)

    try:
        fecha_inicio = datetime.strptime(fecha_inicio_str, '%Y-%m-%d')
        fecha_fin = fecha_inicio + timedelta(days=6)

        inicio = make_aware(datetime.combine(fecha_inicio, datetime.min.time()))
        fin = make_aware(datetime.combine(fecha_fin, datetime.max.time()))

        semana = (
            Promedio.objects
            .filter(id_modelo_id=modelo_id, fecha__range=(inicio, fin))
            .annotate(dia=TruncDate('fecha'))
            .values('dia')
            .annotate(avg_pos=Avg('promedio'), avg_users=Avg('users'))
            .order_by('dia')
        )

        dias = [s['dia'].strftime("%a %d") for s in semana]
        posiciones = [round(s['avg_pos'], 2) for s in semana]
        usuarios = [round(s['avg_users'], 1) for s in semana]

        return JsonResponse({
            'dias': dias,
            'posiciones': posiciones,
            'usuarios': usuarios,
            'fecha_inicio': fecha_inicio_str,
            'fecha_fin': fecha_fin.strftime('%Y-%m-%d')
        })

    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


def obtener_datos_dia(request):
    modelo_id = request.GET.get('modelo')
    fecha_str = request.GET.get('fecha')

    if not (modelo_id and fecha_str):
        return JsonResponse({'error': 'Faltan parámetros'}, status=400)

    try:
        fecha = datetime.strptime(fecha_str, '%Y-%m-%d')
        inicio = make_aware(datetime.combine(fecha, datetime.min.time()))
        fin = make_aware(datetime.combine(fecha, datetime.max.time()))

        registros = (
            Promedio.objects
            .filter(id_modelo_id=modelo_id, fecha__range=(inicio, fin))
            .order_by('fecha')
        )

        horas = [r.fecha.strftime('%H:%M') for r in registros]
        posiciones = [r.promedio for r in registros]
        usuarios = [r.users for r in registros]

        return JsonResponse({
            'horas': horas,
            'posiciones': posiciones,
            'usuarios': usuarios
        })
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

@login_requerido
def graficas_modelo(request):
    id_studio = request.session.get("id_studio")

    # Obtener filtros GET
    modelo_id = request.GET.get("modelo")
    fecha_str = request.GET.get("fecha")

    modelos = ModeloRegistrado.objects.filter(studio_id=id_studio, estado=1)

    registros = []
    horas, posiciones, usuarios = [], [], []

    if modelo_id and fecha_str:
        fecha = datetime.strptime(fecha_str, "%Y-%m-%d")
        inicio = make_aware(datetime.combine(fecha, datetime.min.time()))
        fin = make_aware(datetime.combine(fecha, datetime.max.time()))

        registros = Promedio.objects.filter(
            id_modelo_id=modelo_id,
            fecha__range=(inicio, fin)
        ).order_by('fecha')

        horas = [r.fecha.strftime("%H:%M") for r in registros]
        posiciones = [r.promedio for r in registros]
        usuarios = [r.users for r in registros]

    context = {
        'modelos': modelos,
        'modelo_seleccionado': int(modelo_id) if modelo_id else None,
        'fecha_str': fecha_str,
        'horas': horas,
        'posiciones': posiciones,
        'usuarios': usuarios,
    }

    fecha_inicio_str = request.GET.get('fecha_inicio')
    fecha_fin_str = request.GET.get('fecha_fin')

    dias_semana, promedios_semana, usuarios_semana = [], [], []
    print("Modelo ID:", modelo_id)
    if modelo_id and fecha_inicio_str and fecha_fin_str:
        fecha_inicio = make_aware(datetime.strptime(fecha_inicio_str, "%Y-%m-%d"))
        fecha_fin = make_aware(datetime.strptime(fecha_fin_str, "%Y-%m-%d"))
        
        semana = (
            Promedio.objects
            .filter(id_modelo_id=modelo_id, fecha__range=(fecha_inicio, fecha_fin))
            .annotate(dia=TruncDate('fecha'))
            .values('dia')
            .annotate(avg_pos=Avg('promedio'), avg_users=Avg('users'))
            .order_by('dia')
        )

        dias_semana = [s['dia'].strftime("%a %d") for s in semana]
        promedios_semana = [round(s['avg_pos'], 2) for s in semana]
        usuarios_semana = [round(s['avg_users'], 1) for s in semana]
    
    context.update({
        'fecha_inicio': fecha_inicio_str,
        'fecha_fin': fecha_fin_str,
        'dias_semana': dias_semana,
        'promedios_semana': promedios_semana,
        'usuarios_semana': usuarios_semana,
    })
    return render(request, 'graficas.html', context)

def tokens_por_dia(request):
    modelo_id = request.GET.get('modelo')
    fecha = request.GET.get('fecha')

    tokens = IngresoTokens.objects.filter(id_modelo=modelo_id, fecha=fecha)

    total_cb = sum(t.token_cb for t in tokens)
    total_otro = sum(t.token_otro for t in tokens)

    return JsonResponse({
        'labels': ['Token CB', 'Token Otro'],
        'valores': [total_cb, total_otro]
    })

def tokens_semana_view(request):
    modelo_id = request.GET.get('modelo')
    fecha_inicio = request.GET.get('fecha_inicio')

    if not modelo_id or not fecha_inicio:
        return JsonResponse({'error': 'Faltan parámetros'}, status=400)

    fecha_inicio = datetime.strptime(fecha_inicio, "%Y-%m-%d").date()
    fecha_fin = fecha_inicio + timedelta(days=6)

    tokens = IngresoTokens.objects.filter(
        id_modelo=modelo_id,
        fecha__range=(fecha_inicio, fecha_fin)
    ).aggregate(
        total_cb=Sum('token_cb'),
        total_otro=Sum('token_otro')
    )

    return JsonResponse({
        'labels': ['Token CB', 'Token Otro'],
        'valores': [tokens['total_cb'] or 0, tokens['total_otro'] or 0]
    })

    


@csrf_exempt
def ejecutar_guardar_promedios(request):
    # Autenticación básica por token GET (puedes mejorar esto)
    if request.GET.get("token") != settings.CRON_TOKEN:
        return JsonResponse({"error": "No autorizado"}, status=403)

    call_command("guardar_promedios")
    return JsonResponse({"status": "Comando ejecutado correctamente"})

